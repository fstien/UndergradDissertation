
# coding: utf-8

# In[12]:

from selenium import webdriver
from bs4 import BeautifulSoup
import time
from time import sleep
import timeit
import datetime
import time
import re
import math
import random
import openpyxl
import numpy
import os
import sys
import glob
from sys import exit
import csv
import lxml
import os.path

from xlsxwriter.workbook import Workbook

from difflib import SequenceMatcher

def compare(a, b): return SequenceMatcher(None, a, b).ratio()

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.keys import Keys


# In[13]:

userAgents = [
	"Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.80 Safari/537.36",
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_1) AppleWebKit/601.2.7 (KHTML, like Gecko) Version/9.0.1 Safari/601.2.7",
	"Mozilla/5.0 (Windows NT 6.1; WOW64; rv:41.0) Gecko/20100101 Firefox/41.0",
	"Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.80 Safari/537.36",
	"Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.71 Safari/537.36",
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.80 Safari/537.36",
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.80 Safari/537.36",
	"Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.80 Safari/537.36",
	"Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.86 Safari/537.36",
	"Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; rv:11.0) like Gecko", 
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11) AppleWebKit/601.1.56 (KHTML, like Gecko) Version/9.0 Safari/601.1.56",
	"Mozilla/5.0 (Windows NT 6.1; WOW64; rv:42.0) Gecko/20100101 Firefox/42.0",
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_5) AppleWebKit/601.2.7 (KHTML, like Gecko) Version/9.0.1 Safari/601.2.7",
	"Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.86 Safari/537.36",
	"Mozilla/5.0 (Windows NT 10.0; WOW64; rv:41.0) Gecko/20100101 Firefox/41.0",
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.71 Safari/537.36",
	"Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.71 Safari/537.36",
	"Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:41.0) Gecko/20100101 Firefox/41.0",
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.86 Safari/537.36",
	"Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.80 Safari/537.36",
	"Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.71 Safari/537.36",
	"Mozilla/5.0 (Windows NT 6.3; WOW64; rv:41.0) Gecko/20100101 Firefox/41.0",
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.71 Safari/537.36",
	"Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.80 Safari/537.36",
	"Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.86 Safari/537.36",
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.80 Safari/537.36",
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10.10; rv:41.0) Gecko/20100101 Firefox/41.0",
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.86 Safari/537.36",
	"Mozilla/5.0 (Windows NT 10.0; WOW64; rv:42.0) Gecko/20100101 Firefox/42.0",
	"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/42.0.2311.135 Safari/537.36 Edge/12.10240",
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10.11; rv:41.0) Gecko/20100101 Firefox/41.0",
	"Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:42.0) Gecko/20100101 Firefox/42.0",
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_5) AppleWebKit/601.1.56 (KHTML, like Gecko) Version/9.0 Safari/601.1.56",
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10.11; rv:42.0) Gecko/20100101 Firefox/42.0",
	"Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/45.0.2454.101 Safari/537.36",
	"Mozilla/5.0 (Windows NT 6.3; WOW64; rv:42.0) Gecko/20100101 Firefox/42.0",
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.80 Safari/537.36",
	"Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.71 Safari/537.36"
]


# In[33]:

city = input("Enter city: ")

'''
startValue = input("Start from: ")
endValue = input("End at: ")

startValue = int(startValue)
endValue = int(endValue)
'''

lag = 1

dataFolder = "data/"


# In[ ]:




# In[40]:

def searchData(inputList):
        
    try:
        driver.get("http://www.opentable.com" + inputList[8])
        element = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "review-pagination")))
        time.sleep(0.2)
        bsObj = []
        bsObj.append(BeautifulSoup(driver.page_source, "lxml"))
        
    except: 
        try:
            driver.get("http://www.opentable.com" + inputList[8])
            element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.ID, "review-pagination")))
            time.sleep(0.2)
            bsObj = []
            bsObj.append(BeautifulSoup(driver.page_source, "lxml"))
        except:   
            #print("ERROR: No data found at URL for: " + inputList[2] + ".")   
            return("NoDataFound")
    
    name = inputList[2]
    index = inputList[3]
    cuisine = inputList[5]
    ratingCount = inputList[6]
    price = inputList[7]
    link = inputList[8]
    
    try:
        address = bsObj[0].find(class_="line-height-large").find("div").get_text()
    except:
        #print("ERROR: Could not extract address from page: " + inputList[2] + ".")
        address = "NotFound"
        
    try:
        latitude = bsObj[0].find(class_="line-height-large").find("a")["data-lat"]
        latitude = float(latitude)
        longitude = bsObj[0].find(class_="line-height-large").find("a")["data-long"]
        longitude = float(longitude)
    except:
        #print("ERROR: Could not extract latitude and longitude from page: " + inputList[2] + ".")
        latitude = "NotFound"
        longitude = "NotFound"
    
    try:
        reviewCount = bsObj[0].find(class_="review-count").get_text()
        reviewCount = re.sub("[^0-9]", "", reviewCount)
        reviewCount = int(reviewCount)

        revFreq = len(bsObj[0].findAll(class_="review"))
        pageCount = math.ceil(reviewCount/revFreq)
    except:
        #print("Could not count ratings: " + inputList[2] + ".")
        return("NoDataFound")
    
    try:
        '''
        text = bsObj[0].find(id="profile-details").get_text()
        phonePattern = re.compile(r'(\d{3})\D*(\d{3})\D*(\d{4})\D*(\d*)')
        regexObject = re.search(phonePattern,text)
        numberSplit = regexObject.groups()
        phoneNumber = ""
        for i in numberSplit:  
            phoneNumber += i
        '''
        text = bsObj[0].find(id="profile-details").get_text()
        phoneNumber = text.split("Website")[0].split("Phone Number:")[1].replace("(", "").replace(")", "").replace(" ", "").replace("-", "")
        phoneNumber = int(phoneNumber)
    
    except:
        phoneNumber = "NotFound"
        #print("Could not find phone number: " + inputList[2] + ".")

    try:
        text = bsObj[0].find(id="profile-details").get_text()
        phoneNumber = text.split("Website")[0].split("Phone Number:")[1].replace("(", "").replace(")", "").replace(" ", "").replace("-", "")
        phoneNumber = int(phoneNumber)
    
    except:
        phoneNumber = "NotFound"
        #print("Could not find phone number: " + inputList[2] + ".")
    
    '''
    try: 
        url = bsObj.find(id="profile-details").find(class_="restaurant-website").get_text()[:-1].split("www.")[1]
    except:
        url = "NotFound" 
    '''
        
    '''
    if reviewCount > 2000: 
        pageCount = math.ceil(2000/revFreq)
    '''
    
    for x in range(2, pageCount+1): 

        try:
            time.sleep(round(random.uniform( lag, lag + 1), 2))
            driver.find_element_by_id("PageLink" + str(x)).click()
            element = WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.ID, "reviews-page")))
            time.sleep(0.1)
            bsObj.append(BeautifulSoup(driver.page_source, "lxml"))
        except:
            try:
                time.sleep(round(random.uniform( lag, lag + 2), 2))
                driver.find_element_by_id("PageLink" + str(x)).click()
                element = WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.ID, "reviews-page")))
                time.sleep(0.1)
                bsObj.append(BeautifulSoup(driver.page_source, "lxml"))
            except:
                try:
                    time.sleep(round(random.uniform( lag, lag + 3), 2))
                    driver.find_element_by_id("PageLink" + str(x)).click()
                    element = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.ID, "reviews-page")))
                    time.sleep(0.1)
                    bsObj.append(BeautifulSoup(driver.page_source, "lxml"))
                except:
                    #print("ERROR: Ajax loop for " + inputList[2] + " at " + str(x) + "/" + str(pageCount+1))
                    return("NoDataFound")

    bsObj = list(set(bsObj))
    
    ratings = []

    try:
        for bs in bsObj: 
            reviews = bs.findAll(class_="review")

            for review in reviews:
                stars = review.find(class_="all-stars filled")["title"]
                ratings.append(int(stars))
    except:
        #print("ERROR: Loop error for ratings extraction: " + inputList[2] + ".")
        return("NoDataFound")

    try:
        meanRating = numpy.mean(ratings)
        
        meanRating = round(meanRating, 4)
        
        reviewCount = len(ratings)

        histogram = {
            "1": ratings.count(1),
            "2": ratings.count(2),
            "3": ratings.count(3),
            "4": ratings.count(4),
            "5": ratings.count(5),
        }

    except:
        #print("ERROR: Could not compute ratings data." + inputList[2] + ".")
        return("NoDataFound")
    
    try:
        
        dataTime = time.strftime("%a, %d %b %Y %X", time.gmtime()).replace(",", "")
        
        dataList = [dataTime, city, name, index, ratingCount, price, cuisine, longitude, latitude, address, phoneNumber, meanRating, histogram["1"], histogram["2"], histogram["3"], histogram["4"], histogram["5"], link]
    
        return(dataList)
    
    except:
        #print("ERROR: Could not build dataList: " + inputList[2] + ".")
        return("NoDataFound")



# In[35]:

def writeHeader():
    with open(dataFolder + city + "OTData.csv", "a") as csvFile: 
        wr = csv.writer(csvFile)
        wr.writerow(["Time", "City", "Name", "Index", "ReviewCount", "Price", "Cuisine", "Longitude", "Latitude", "Address", "PhoneNumber", "MeanRating", "1", "2", "3", "4", "5", "Link"])

if os.path.isfile(dataFolder + city + "OTData.csv"): 
    if os.stat(dataFolder + city + "OTData.csv").st_size == 0:
        writeHeader()
else:
    writeHeader()


# In[14]:

randomUA = random.choice(userAgents)
dcap = dict(DesiredCapabilities.PHANTOMJS)
dcap["phantomjs.page.settings.userAgent"] = randomUA
driver = webdriver.PhantomJS(desired_capabilities=dcap)
driver.set_window_size(1100 + random.randint(-100, 100), 1000 + random.randint(-100, 100))


# In[37]:

def printData(data):
    with open(dataFolder + city + "OTData.csv", "a") as csvFile: 
        wr = csv.writer(csvFile)
        wr.writerow(data)


# In[38]:

filename = dataFolder + city + "OTLinks"

# THE CONTENT OF THE LOADWORKBOOK FUNCTION
while True: 
    try:
        wb = openpyxl.load_workbook(filename + '.xlsx')
        sheet = wb.get_sheet_by_name('Sheet1')
        break
    except:
        time.sleep(random.random()/2 + 0.1)
        pass

column_count = sheet.max_row

textFile = city + "OTLinks"

if len(glob.glob(textFile + "*")) == 0:
    open(textFile + "=READY.txt", "w")

    
def loadWorkbook():
    global wb
    global sheet
    while True: 
        try:
            wb = openpyxl.load_workbook(filename + '.xlsx')
            sheet = wb.get_sheet_by_name('Sheet1')
            break
        except:
            time.sleep(random.random()/2 + 0.5)
            pass

def rowsRemaining():
    loadWorkbook()
    for i in range(2, column_count + 1):
        if str(sheet.cell(row=i, column=5).value) == "0":
            return(True)
    return(False)   

def saveNewStatus(rowNumer, status):
    while True:
        try:
            while glob.glob(textFile + "=*") == textFile + "=BUSY.txt":
                sleep(2*random.random())
            os.rename(textFile + "=READY.txt", textFile + "=BUSY.txt")
            
            wb = openpyxl.load_workbook(filename + '.xlsx')
            sheet = wb.get_sheet_by_name('Sheet1')
            sheet.cell(row=rowNumer, column=5).value = str(status)
            wb.save(filename + '.xlsx')
            
            os.rename(textFile + "=BUSY.txt", textFile + "=READY.txt")
            break
        except:
            time.sleep(random.random()/2 + 0.5)
            pass

def randomRow():
    loadWorkbook()
    while True:
        randNum = random.randrange(column_count - 1) + 2
        if str(sheet.cell(row=randNum, column=5).value) == str(0):
            
            saveNewStatus(randNum, 1)
            
            row = []
            for i in range(1, 10):
                row.append(sheet.cell(row=randNum, column=i).value)
            return(row)

        
while rowsRemaining():
    
    row = randomRow()
    
    data = []
    data = searchData(row) 
        
    if data == "NoDataFound":
        print("ERROR: Data not found for " + row[2] + ".")
        saveNewStatus((int(row[3]) + 1), "NoDataFound")
    else:
        print("SUCCESS: Data found for " + row[2] + ".")
        printData(data)
    
    if random.randint(0, 1)>0.9:
        driver.quit()
        time.sleep(5)
        while True:
            try:
                randomUA = random.choice(userAgents)
                dcap = dict(DesiredCapabilities.PHANTOMJS)
                dcap["phantomjs.page.settings.userAgent"] = randomUA
                driver = webdriver.PhantomJS(desired_capabilities=dcap)
                driver.set_window_size(1100 + random.randint(-100, 100), 1000 + random.randint(-100, 100))
                break
            except:
                time.sleep(5)
                pass
        
    
driver.quit()
try:
    os.remove(textFile + "=READY.txt")
except:
    pass


# In[ ]:




# In[ ]:




# In[ ]:




# In[ ]:




# In[ ]:




# In[ ]:




# In[ ]:




# In[ ]:




# In[ ]:




# In[5]:

'''

with open(dataFolder + city + "OTLinks.csv") as csv_file:
    
    next(csv_file, None) 
    
    for index, row in enumerate(csv.reader(csv_file, delimiter=',')):

        if index >= startValue and index <= endValue:
            
            getData(row)        
            
            if random.randint(0, 1)>0.95:
                driver.quit()
                time.sleep(10)
                randomUA = random.choice(userAgents)
                dcap = dict(DesiredCapabilities.PHANTOMJS)
                dcap["phantomjs.page.settings.userAgent"] = randomUA
                driver = webdriver.PhantomJS(desired_capabilities=dcap)
                driver.set_window_size(1100 + random.randint(-100, 100), 1000 + random.randint(-100, 100))
    
driver.quit()


'''


# In[ ]:




# In[ ]:




# In[ ]:




# In[ ]:




# In[13]:

'''
with open(filename + ".csv") as f:
    reader = csv.reader(f,delimiter = ",")
    data = list(reader)
    row_count = len(data) - 1
    
with open(filename + ".csv") as csv_file:

    next(csv_file, None)
    
    fileObject = csv.reader(csv_file, delimiter = ",")
    
    def rowsRemaining():
        for index, row in enumerate(fileObject):
            if int(row[3]) == 0:
                return(True)
        return(False)
    
    def editStatus(index, status):
        time.sleep(1)
        os.rename(filename + ".csv", filename + "BUSY" + ".csv")
        time.sleep(1)
        r = csv.reader(open(filename + "BUSY" + ".csv"))
        lines = [l for l in r]
        lines[index][3] = status
        writer = csv.writer(open(filename + "BUSY" + ".csv", 'w'))
        writer.writerows(lines)
        time.sleep(1)
        os.rename(filename + "BUSY" + ".csv", filename + ".csv")
        time.sleep(1)
        
    def randomFile():
        #if rowsRemaining():
        if True: 
            randNum = random.randrange(row_count)
            for index, row in enumerate(fileObject):
                if index == randNum:
                    #editStatus(index + 1, 1)
                    time.sleep(1)
                    print(row)
                    return(row)
        else:
            return("NotFound")
        
    
    for i in range(1, 10):
        
        time.sleep(1)
        print(randomFile())
       
        #if row != "NotFound":
            
            
'''



# In[ ]:




# In[ ]:




# In[ ]:




# In[ ]:




# In[ ]:




# In[ ]:



