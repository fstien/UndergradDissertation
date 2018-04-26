
# coding: utf-8

# In[1]:

from selenium import webdriver
from bs4 import BeautifulSoup
import time
from time import sleep
import timeit
import datetime
import time
import re
import math
import random
import openpyxl
import numpy
import os
import sys
import glob
from sys import exit
import csv
import lxml
import os.path

from xlsxwriter.workbook import Workbook

from difflib import SequenceMatcher

def compare(a, b): return SequenceMatcher(None, a, b).ratio()

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.keys import Keys


# In[2]:

userAgents = [
	"Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.80 Safari/537.36",
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_1) AppleWebKit/601.2.7 (KHTML, like Gecko) Version/9.0.1 Safari/601.2.7",
	"Mozilla/5.0 (Windows NT 6.1; WOW64; rv:41.0) Gecko/20100101 Firefox/41.0",
	"Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.80 Safari/537.36",
	"Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.71 Safari/537.36",
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.80 Safari/537.36",
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.80 Safari/537.36",
	"Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.80 Safari/537.36",
	"Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.86 Safari/537.36",
	"Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; rv:11.0) like Gecko", 
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11) AppleWebKit/601.1.56 (KHTML, like Gecko) Version/9.0 Safari/601.1.56",
	"Mozilla/5.0 (Windows NT 6.1; WOW64; rv:42.0) Gecko/20100101 Firefox/42.0",
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_5) AppleWebKit/601.2.7 (KHTML, like Gecko) Version/9.0.1 Safari/601.2.7",
	"Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.86 Safari/537.36",
	"Mozilla/5.0 (Windows NT 10.0; WOW64; rv:41.0) Gecko/20100101 Firefox/41.0",
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.71 Safari/537.36",
	"Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.71 Safari/537.36",
	"Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:41.0) Gecko/20100101 Firefox/41.0",
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.86 Safari/537.36",
	"Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.80 Safari/537.36",
	"Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.71 Safari/537.36",
	"Mozilla/5.0 (Windows NT 6.3; WOW64; rv:41.0) Gecko/20100101 Firefox/41.0",
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.71 Safari/537.36",
	"Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.80 Safari/537.36",
	"Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.86 Safari/537.36",
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.80 Safari/537.36",
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10.10; rv:41.0) Gecko/20100101 Firefox/41.0",
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.86 Safari/537.36",
	"Mozilla/5.0 (Windows NT 10.0; WOW64; rv:42.0) Gecko/20100101 Firefox/42.0",
	"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/42.0.2311.135 Safari/537.36 Edge/12.10240",
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10.11; rv:41.0) Gecko/20100101 Firefox/41.0",
	"Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:42.0) Gecko/20100101 Firefox/42.0",
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_5) AppleWebKit/601.1.56 (KHTML, like Gecko) Version/9.0 Safari/601.1.56",
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10.11; rv:42.0) Gecko/20100101 Firefox/42.0",
	"Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/45.0.2454.101 Safari/537.36",
	"Mozilla/5.0 (Windows NT 6.3; WOW64; rv:42.0) Gecko/20100101 Firefox/42.0",
	"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.80 Safari/537.36",
	"Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2490.71 Safari/537.36"
]


# In[56]:

city = input("Enter city: ")

'''
startValue = input("Start from: ")
endValue = input("End at: ")

startValue = int(startValue)
endValue = int(endValue)
'''

lag = 1

dataFolder = "data/"


# In[ ]:




# In[57]:

def searchData(inputList):
        
    try:
        driver.get("https://www.tripadvisor.com" + inputList[9])
        time.sleep(random.randint(lag, lag + 1))    
        bsObj = BeautifulSoup(driver.page_source, "lxml")
    except: 
        print("ERROR: Could not load page: " + inputList[2] + ".")   
        return("NoDataFound")
    
    name = inputList[2]
    ranking = inputList[3]
    cuisine = inputList[5]
    longitude = inputList[6]
    latitude = inputList[7]
    reviewCount = inputList[8]
    link = inputList[9]
    
    try:
        price = bsObj.find(id="HEADING_GROUP").find(class_="price_rating").get_text()
        price = price.replace("\n", " ").replace(" ", "")
        price = len(price)
    except:
        #print("ERROR: Could not find price: " + inputList[2] + ".")
        price = "NotFound"
    
    try:
        address = bsObj.find(id="ABOVE_THE_FOLD").find("address").get_text()
        address = address.replace("\n", "")
    except:
        #print("ERROR: Could not find address: " + inputList[2] + ".")  
        address = "NotFound"
        
    try: 
        phoneNumber = bsObj.find(class_="phoneNumber").get_text()
        phoneNumber = phoneNumber.replace(" ", "").replace("-", "").replace("+", "").replace("(", "").replace(")", "")
    except:
        #print("ERROR: Could not find phone number: " + inputList[2] + ".")  
        phoneNumber = "NotFound"
        
    try:
        ratings = []
        ratingList = bsObj.find(id="ratingFilter").find("ul").findAll("li")
        for i in range(0, len(ratingList)):
            count = ratingList[i].get_text().replace(",", "").replace(" ", "").replace("\n", "").split("(")[1].split(")")[0]
            ratings.extend([(5-i)]*int(count))
    except:
        #print("ERROR: Could not find the ratings: " + inputList[2] + ".")  
        ratings = "NotFound"

    try:
        meanRating = numpy.mean(ratings)
        meanRating = round(meanRating, 4)
        histogram = {
            "1": ratings.count(1),
            "2": ratings.count(2),
            "3": ratings.count(3),
            "4": ratings.count(4),
            "5": ratings.count(5),
        }
    except: 
        #print("ERROR: Could not compute the ratings statistics: "  + inputList[2] + ".")  
        histogram = {}
        meanRating = histogram["1"] = histogram["2"] = histogram["3"] = histogram["4"] =histogram["5"] = "NotFound"
    
    try:
        dataTime = time.strftime("%a, %d %b %Y %X", time.gmtime()).replace(",", "")
        dataList = [dataTime, city, name, reviewCount, price, cuisine, longitude, latitude, address, phoneNumber, ranking, meanRating, histogram["1"], histogram["2"], histogram["3"], histogram["4"], histogram["5"], link]
        return(dataList)
    except:
        #print("ERROR: Could not build dataList: "  + inputList[2] + ".")  
        return("NoDataFound")



# In[58]:

def topRowWrite():
    with open(dataFolder + city + "TAData.csv", "a") as csvFile: 
        wr = csv.writer(csvFile)
        wr.writerow(["Time", "City", "Name", "ReviewCount", "Price", "Cuisine", "Longitude", "Latitude", "Address", "PhoneNumber", "Ranking", "MeanRating", "1", "2", "3", "4", "5", "Link"])

if os.path.isfile(dataFolder + city + "TAData.csv"): 
    if os.stat(dataFolder + city + "TAData.csv").st_size == 0:
        topRowWrite()
else:
    topRowWrite()


# In[3]:

randomUA = random.choice(userAgents)
dcap = dict(DesiredCapabilities.PHANTOMJS)
dcap["phantomjs.page.settings.userAgent"] = randomUA
driver = webdriver.PhantomJS(desired_capabilities=dcap)
driver.set_window_size(1100 + random.randint(-100, 100), 1000 + random.randint(-100, 100))


# In[57]:

def printData(data):
    with open(dataFolder + city + "TAData.csv", "a") as csvFile: 
        wr = csv.writer(csvFile)
        wr.writerow(data)


# In[61]:

filename = dataFolder + city + "TALinks"

# THE CONTENT OF THE LOADWORKBOOK FUNCTION
while True: 
    try:
        wb = openpyxl.load_workbook(filename + '.xlsx')
        sheet = wb.get_sheet_by_name('Sheet1')
        break
    except:
        time.sleep(random.random()/2 + 0.1)
        pass
        
column_count = sheet.max_row

textFile = city + "TALinks"

if len(glob.glob(textFile + "*")) == 0:
    open(textFile + "=READY.txt", "w")


def loadWorkbook():
    global wb
    global sheet
    while True: 
        try:
            wb = openpyxl.load_workbook(filename + '.xlsx')
            sheet = wb.get_sheet_by_name('Sheet1')
            break
        except:
            time.sleep(random.random()/2 + 0.5)
            pass

def rowsRemaining():
    loadWorkbook()
    for i in range(2, column_count):
        if str(sheet.cell(row=i, column=5).value) == str(0):
            return(True)
    return(False)  

def saveNewStatus(rowNumer, status):
    while True:
        try:
            while glob.glob(textFile + "=*") == textFile + "=BUSY.txt":
                sleep(2*random.random())
            os.rename(textFile + "=READY.txt", textFile + "=BUSY.txt")
            
            wb = openpyxl.load_workbook(filename + '.xlsx')
            sheet = wb.get_sheet_by_name('Sheet1')
            sheet.cell(row=rowNumer, column=5).value = str(status)
            wb.save(filename + '.xlsx')
            
            os.rename(textFile + "=BUSY.txt", textFile + "=READY.txt")
            break
        except:
            time.sleep(random.random()/2 + 0.5)
            pass
        
def randomRow():
    loadWorkbook()
    while True:
        randNum = random.randrange(column_count - 1) + 2
        if str(sheet.cell(row=randNum, column=5).value) == str(0):
            
            saveNewStatus(randNum, 1)
            
            row = []
            for i in range(1, 11):
                row.append(sheet.cell(row=randNum, column=i).value)
            return(row)
        

while rowsRemaining():
    
    row = randomRow()
    
    data = []
    data = searchData(row) 
        
    if data == "NoDataFound":
        print("ERROR: Data not found for " + row[2] + ".")
        saveNewStatus((int(row[3]) + 1), "NoDataFound")
    else:
        print("SUCCESS: Data found for " + row[2] + ".")
        printData(data)
    
    if random.randint(0, 1)>0.95:
        driver.quit()
        time.sleep(5)
        while True:
            try:
                randomUA = random.choice(userAgents)
                dcap = dict(DesiredCapabilities.PHANTOMJS)
                dcap["phantomjs.page.settings.userAgent"] = randomUA
                driver = webdriver.PhantomJS(desired_capabilities=dcap)
                driver.set_window_size(1100 + random.randint(-100, 100), 1000 + random.randint(-100, 100))
                break
            except:
                time.sleep(5)
                pass
        

driver.quit()

os.remove(textFile + "=READY.txt")


# In[ ]:




# In[ ]:




# In[64]:




# In[44]:

'''
driver.get("https://www.tripadvisor.co.uk/Restaurant_Review-g36471-d908858-Reviews-George_s_Restaurant-Oak_Park_Illinois.html")
bsObj = BeautifulSoup(driver.page_source, "lxml")

linkArray = bsObj.findAll(class_="taLnk")
for index, item in enumerate(linkArray):
    if item.get_text() == "Website":
        linkIndex = index

driver.find_element_by_xpath("//span[@class='taLnk' and text()='Website']").click()
time.sleep(5)
driver.current_url
'''


# In[ ]:




# In[13]:

'''
with open(dataFolder + city + "TALinks.csv") as csv_file:
    
    next(csv_file, None) 

    for index, row in enumerate(csv.reader(csv_file, delimiter=',')):
        
        if index >= startValue and index <= endValue:
                        
            searchData(row)

            if random.randint(0, 1)>0.95:
                driver.quit()
                time.sleep(10)    
                randomUA = random.choice(userAgents)
                dcap = dict(DesiredCapabilities.PHANTOMJS)
                dcap["phantomjs.page.settings.userAgent"] = randomUA
                driver = webdriver.PhantomJS(desired_capabilities=dcap)
                driver.set_window_size(1100 + random.randint(-100, 100), 1000 + random.randint(-100, 100))
        
driver.quit()
'''


# In[ ]:

'''
    
    try:
        bsObj = BeautifulSoup(driver.page_source, "lxml")

        link = driver.current_url
        link = link.replace("\n", " ").replace(" ", "")
    
    except: 
        print("ERROR: Unexpected error in finding link.")

    try:
        name = bsObj.find(id="HEADING").get_text()
        name = name.replace("\n", "")
    except:
        print("ERROR: Could not find name: " + inputList[3] + ".")
        name = "NotFound"

    try: 
        reviewCount = bsObj.find(id="TABS_REVIEWS").find(class_="tabs_pers_counts").get_text()
        reviewCount = re.sub("[^0-9]", "", reviewCount)
        reviewCount = reviewCount.replace("\n", "").replace(" ", "")
        reviewCount = int(reviewCount)
    except:
        print("ERROR: Could not find number of reviews.")
        reviewCount = "NotFound"	

    try:
        ranking = bsObj.find(id="HEADING_GROUP").find(class_="slim_ranking").get_text()
        ranking = ranking.replace("\n","").split(" ")[0].replace("#", "").replace(",", "")
    except:
        print("ERROR: Could not find number of reviews.")
        reviewCount = "NotFound"


    try:
        cuisine = bsObj.find(class_="heading_details").find("a").get_text()
        cuisine = cuisine.replace("\n", " ").replace(" ", "")
    except:
        print("ERROR: Could not find cuisine.")
        cuisine = "NotFound"

    try:
        longitude = bsObj.find(id="NEARBY_TAB").find(class_="mapWrap").find(class_="mapContainer")["data-lng"]
        longitude = float(longitude)
    except:
        print("ERROR: Could not find longitude.")
        longitude = "NotFound"

    try:
        latitude = bsObj.find(id="NEARBY_TAB").find(class_="mapWrap").find(class_="mapContainer")["data-lat"]
        latitude = float(latitude)
    except:
        print("ERROR: Could not find latitude.")
        latitude = "NotFound"



'''


# In[ ]:



